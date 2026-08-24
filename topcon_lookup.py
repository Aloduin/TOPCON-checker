"""Look up a small set of patient IDs in a full TOPCON data copy.

This module provides both a reusable Python API and a command-line interface.
The graphical interface in ``topcon_lookup_gui.py`` uses the same functions.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from parse_topcon import (
    ImageRecord,
    PatientRecord,
    TopconFormatError,
    parse_history,
    parse_patients,
)


IMAGE_SUFFIXES = {".JPG", ".JPEG", ".PNG", ".BMP", ".TIF", ".TIFF"}
HEADER_ALIASES = {
    "id",
    "patientid",
    "patient_id",
    "患者id",
    "患者号",
    "病历号",
    "就诊号",
}

STATUS_COMPLETE = "患者及图片均已找到"
STATUS_INDEX_ONLY = "仅找到图片索引"
STATUS_PARTIAL = "部分图片文件缺失"
STATUS_DUPLICATE_FILES = "图片文件存在重名"
STATUS_NO_INDEX = "患者已找到，无图片索引"
STATUS_NOT_FOUND = "未找到患者"
STATUS_AMBIGUOUS = "患者号匹配不唯一"

FILE_FOUND = "文件已找到"
FILE_MISSING = "文件缺失"
FILE_DUPLICATE = "同名文件多个"


@dataclass(frozen=True, slots=True)
class DatasetFiles:
    root: Path
    metadata_dir: Path
    patients_file: Path
    history_file: Path
    label_file: Path | None


@dataclass(frozen=True, slots=True)
class QuerySummary:
    query_order: int
    requested_id: str
    match_status: str
    match_mode: str
    patient_id: str
    patient_name: str
    sex: str
    birth_date: str
    image_index_count: int
    image_files_found: int
    image_files_missing: int
    latest_capture_date: str
    notes: str


@dataclass(frozen=True, slots=True)
class ImageLookupRow:
    query_order: int
    requested_id: str
    match_mode: str
    patient_id: str
    patient_name: str
    sex: str
    birth_date: str
    registered_date: str
    capture_date: str
    capture_time: str
    capture_source: str
    image_number_in_exam: str
    image_filename: str
    file_status: str
    image_path: str
    record_index: int


@dataclass(frozen=True, slots=True)
class LookupResult:
    dataset: DatasetFiles
    image_root: Path
    generated_at: datetime
    requested_ids: tuple[str, ...]
    summaries: tuple[QuerySummary, ...]
    images: tuple[ImageLookupRow, ...]

    @property
    def stats(self) -> dict[str, int]:
        return {
            "requested_ids": len(self.requested_ids),
            "matched_patients": sum(
                row.match_status not in {STATUS_NOT_FOUND, STATUS_AMBIGUOUS}
                for row in self.summaries
            ),
            "patients_with_image_index": sum(
                row.image_index_count > 0 for row in self.summaries
            ),
            "image_records": len(self.images),
            "image_files_found": sum(
                row.file_status in {FILE_FOUND, FILE_DUPLICATE} for row in self.images
            ),
            "image_files_missing": sum(
                row.file_status == FILE_MISSING for row in self.images
            ),
            "queries_requiring_attention": sum(
                row.match_status != STATUS_COMPLETE for row in self.summaries
            ),
        }


@dataclass(frozen=True, slots=True)
class CopyRecord:
    query_order: int
    requested_id: str
    patient_id: str
    image_filename: str
    source_path: str
    copy_status: str
    destination_path: str
    error: str


@dataclass(frozen=True, slots=True)
class CopyResult:
    destination_root: Path
    records: tuple[CopyRecord, ...]

    @property
    def stats(self) -> dict[str, int]:
        return {
            "copied": sum(row.copy_status == "已复制" for row in self.records),
            "already_present": sum(
                row.copy_status == "已存在（内容相同）" for row in self.records
            ),
            "renamed_conflicts": sum(
                row.copy_status == "已复制（同名改名）" for row in self.records
            ),
            "failed": sum(row.copy_status == "拷贝失败" for row in self.records),
        }


def normalize_ids(values: Iterable[object]) -> list[str]:
    """Normalize ID values while preserving their order and leading zeroes."""

    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value is None or pd.isna(value):
            continue
        patient_id = str(value).strip()
        if re.fullmatch(r"[0-9]+\.0", patient_id):
            patient_id = patient_id[:-2]
        if not patient_id or patient_id.lower().replace(" ", "") in HEADER_ALIASES:
            continue
        if patient_id not in seen:
            seen.add(patient_id)
            result.append(patient_id)
    return result


def parse_id_text(text: str) -> list[str]:
    return normalize_ids(re.split(r"[\s,，;；]+", text))


def load_ids_file(path: str | Path) -> list[str]:
    """Load IDs from the first non-empty column of TXT, CSV, or XLSX."""

    input_path = Path(path)
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        frame = pd.read_excel(input_path, header=None, dtype=str, keep_default_na=False)
    elif suffix == ".csv":
        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                frame = pd.read_csv(
                    input_path,
                    header=None,
                    dtype=str,
                    keep_default_na=False,
                    encoding=encoding,
                    sep=None,
                    engine="python",
                )
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        else:
            raise ValueError("CSV 不是有效的 UTF-8 或 GB18030 文件") from last_error
    elif suffix == ".txt":
        raw = input_path.read_bytes()
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                return parse_id_text(raw.decode(encoding))
            except UnicodeDecodeError:
                continue
        raise ValueError("TXT 不是有效的 UTF-8 或 GB18030 文件")
    else:
        raise ValueError("仅支持 TXT、CSV、XLSX 和 XLSM ID 文件")

    non_empty_columns = [
        column
        for column in frame.columns
        if frame[column].astype(str).str.strip().ne("").any()
    ]
    if not non_empty_columns:
        return []
    return normalize_ids(frame[non_empty_columns[0]].tolist())


def locate_dataset(root: str | Path) -> DatasetFiles:
    """Locate one HIST0010 dataset below a full-copy directory."""

    root_path = Path(root).resolve()
    if not root_path.is_dir():
        raise TopconFormatError(f"数据目录不存在：{root_path}")
    history_candidates = sorted(
        path for path in root_path.rglob("*") if path.is_file() and path.name.upper() == "HIST0010"
    )
    if len(history_candidates) != 1:
        raise TopconFormatError(
            f"数据目录中找到 {len(history_candidates)} 个 HIST0010；请直接选择包含目标数据集的目录"
        )
    history_file = history_candidates[0]
    metadata_dir = history_file.parent
    patient_candidates = sorted(
        path for path in metadata_dir.iterdir() if path.is_file() and path.suffix.lower() == ".txt"
    )
    if len(patient_candidates) != 1:
        raise TopconFormatError(
            f"{metadata_dir} 中找到 {len(patient_candidates)} 个患者 TXT 文件，无法自动确定"
        )
    label_path = metadata_dir / "LABEL"
    return DatasetFiles(
        root=root_path,
        metadata_dir=metadata_dir,
        patients_file=patient_candidates[0],
        history_file=history_file,
        label_file=label_path if label_path.is_file() else None,
    )


def index_image_files(root: str | Path) -> dict[str, list[Path]]:
    """Build a case-insensitive filename index without modifying copied data."""

    root_path = Path(root).resolve()
    if not root_path.is_dir():
        raise TopconFormatError(f"图片目录不存在：{root_path}")
    indexed: dict[str, list[Path]] = defaultdict(list)
    skipped_dirs = {".git", ".venv", ".uv-cache", "__pycache__", "output"}
    for current_root, dir_names, file_names in os.walk(root_path):
        dir_names[:] = [
            name
            for name in dir_names
            if name.lower() not in skipped_dirs
            and not name.startswith(".")
            and not name.startswith("TOPCON核查_")
        ]
        current_path = Path(current_root)
        for file_name in file_names:
            if Path(file_name).suffix.upper() in IMAGE_SUFFIXES:
                indexed[file_name.upper()].append((current_path / file_name).resolve())
    return dict(indexed)


def _canonical_numeric_id(value: str) -> str | None:
    return str(int(value)) if value.isdigit() else None


def _resolve_patient_id(
    requested_id: str,
    available_ids: set[str],
    canonical_ids: dict[str, set[str]],
    allow_leading_zero_match: bool,
) -> tuple[str | None, str, bool]:
    if requested_id in available_ids:
        return requested_id, "精确匹配", False
    canonical = _canonical_numeric_id(requested_id)
    if not allow_leading_zero_match or canonical is None:
        return None, "未匹配", False
    candidates = canonical_ids.get(canonical, set())
    if len(candidates) == 1:
        return next(iter(candidates)), "忽略前导零（唯一匹配）", False
    return None, "忽略前导零匹配不唯一" if candidates else "未匹配", len(candidates) > 1


def lookup_patient_ids(
    data_root: str | Path,
    requested_ids: Iterable[object],
    *,
    image_root: str | Path | None = None,
    allow_leading_zero_match: bool = False,
) -> LookupResult:
    """Resolve requested IDs to patients, image metadata, and actual image paths."""

    normalized_ids = normalize_ids(requested_ids)
    if not normalized_ids:
        raise ValueError("没有可查询的患者 ID")

    dataset = locate_dataset(data_root)
    patients = parse_patients(dataset.patients_file)
    history = parse_history(dataset.history_file)
    resolved_image_root = Path(image_root).resolve() if image_root else dataset.root
    image_files = index_image_files(resolved_image_root)

    patients_by_id: dict[str, PatientRecord] = {}
    for patient in patients:
        patients_by_id.setdefault(patient.patient_id, patient)
    history_by_id: dict[str, list[ImageRecord]] = defaultdict(list)
    for image in history:
        history_by_id[image.patient_id].append(image)

    available_ids = set(patients_by_id) | set(history_by_id)
    canonical_ids: dict[str, set[str]] = defaultdict(set)
    for patient_id in available_ids:
        canonical = _canonical_numeric_id(patient_id)
        if canonical is not None:
            canonical_ids[canonical].add(patient_id)

    summaries: list[QuerySummary] = []
    image_rows: list[ImageLookupRow] = []
    for query_order, requested_id in enumerate(normalized_ids, start=1):
        actual_id, match_mode, ambiguous = _resolve_patient_id(
            requested_id,
            available_ids,
            canonical_ids,
            allow_leading_zero_match,
        )
        if actual_id is None:
            summaries.append(
                QuerySummary(
                    query_order=query_order,
                    requested_id=requested_id,
                    match_status=STATUS_AMBIGUOUS if ambiguous else STATUS_NOT_FOUND,
                    match_mode=match_mode,
                    patient_id="",
                    patient_name="",
                    sex="",
                    birth_date="",
                    image_index_count=0,
                    image_files_found=0,
                    image_files_missing=0,
                    latest_capture_date="",
                    notes="请核对患者号原始格式" if ambiguous else "患者清单和 HIST0010 中均未找到",
                )
            )
            continue

        patient = patients_by_id.get(actual_id)
        indexed_images = sorted(
            history_by_id.get(actual_id, []),
            key=lambda row: (row.capture_date, row.capture_time, row.record_index),
        )
        representative = indexed_images[0] if indexed_images else None
        patient_name = patient.patient_name if patient else representative.patient_name
        sex = patient.sex if patient else representative.sex
        birth_date = patient.birth_date if patient else representative.birth_date
        registered_date = patient.registered_date if patient else representative.registered_date

        found_count = 0
        missing_count = 0
        duplicate_count = 0
        for image in indexed_images:
            paths = image_files.get(image.image_filename.upper(), [])
            if not paths:
                file_status = FILE_MISSING
                image_path = ""
                missing_count += 1
            elif len(paths) == 1:
                file_status = FILE_FOUND
                image_path = str(paths[0])
                found_count += 1
            else:
                file_status = FILE_DUPLICATE
                image_path = "\n".join(str(path) for path in paths)
                found_count += 1
                duplicate_count += 1
            image_rows.append(
                ImageLookupRow(
                    query_order=query_order,
                    requested_id=requested_id,
                    match_mode=match_mode,
                    patient_id=actual_id,
                    patient_name=patient_name,
                    sex=sex,
                    birth_date=birth_date,
                    registered_date=registered_date,
                    capture_date=image.capture_date,
                    capture_time=image.capture_time,
                    capture_source=image.capture_source,
                    image_number_in_exam=image.image_number_in_exam,
                    image_filename=image.image_filename,
                    file_status=file_status,
                    image_path=image_path,
                    record_index=image.record_index,
                )
            )

        if not indexed_images:
            match_status = STATUS_NO_INDEX
        elif duplicate_count:
            match_status = STATUS_DUPLICATE_FILES
        elif missing_count == len(indexed_images):
            match_status = STATUS_INDEX_ONLY
        elif missing_count:
            match_status = STATUS_PARTIAL
        else:
            match_status = STATUS_COMPLETE

        notes: list[str] = []
        if patient is None:
            notes.append("仅在 HIST0010 中找到患者信息")
        if match_mode != "精确匹配":
            notes.append(match_mode)
        if duplicate_count:
            notes.append(f"{duplicate_count} 个图片名对应多个文件")
        summaries.append(
            QuerySummary(
                query_order=query_order,
                requested_id=requested_id,
                match_status=match_status,
                match_mode=match_mode,
                patient_id=actual_id,
                patient_name=patient_name,
                sex=sex,
                birth_date=birth_date,
                image_index_count=len(indexed_images),
                image_files_found=found_count,
                image_files_missing=missing_count,
                latest_capture_date=max(
                    (image.capture_date for image in indexed_images), default=""
                ),
                notes="；".join(notes),
            )
        )

    return LookupResult(
        dataset=dataset,
        image_root=resolved_image_root,
        generated_at=datetime.now().astimezone(),
        requested_ids=tuple(normalized_ids),
        summaries=tuple(summaries),
        images=tuple(image_rows),
    )


def _safe_path_component(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip(". ")
    if cleaned:
        return cleaned
    return "patient_" + hashlib.sha256(value.encode("utf-8")).hexdigest()[:12]


def _folder_sex(value: str) -> str:
    normalized = value.strip().lower()
    if normalized in {"m", "male", "男"}:
        return "男"
    if normalized in {"f", "female", "女"}:
        return "女"
    return value.strip() or "性别未知"


def _patient_folder_name(image: ImageLookupRow) -> str:
    patient_id = _safe_path_component(image.patient_id or image.requested_id)
    birth_date = _safe_path_component(image.birth_date or "出生日期未知")
    sex = _safe_path_component(_folder_sex(image.sex))
    return f"{patient_id}_{birth_date}_{sex}"


def _same_file_contents(first: Path, second: Path) -> bool:
    if first.stat().st_size != second.stat().st_size:
        return False
    first_hash = hashlib.sha256()
    second_hash = hashlib.sha256()
    with first.open("rb") as first_handle, second.open("rb") as second_handle:
        while True:
            first_chunk = first_handle.read(1024 * 1024)
            second_chunk = second_handle.read(1024 * 1024)
            if not first_chunk and not second_chunk:
                break
            first_hash.update(first_chunk)
            second_hash.update(second_chunk)
    return first_hash.digest() == second_hash.digest()


def _available_destination(source: Path, patient_dir: Path) -> tuple[Path, str]:
    preferred = patient_dir / source.name
    if not preferred.exists():
        return preferred, "已复制"
    if preferred.is_file() and _same_file_contents(source, preferred):
        return preferred, "已存在（内容相同）"
    for number in range(2, 10_000):
        candidate = patient_dir / f"{source.stem}__{number}{source.suffix}"
        if not candidate.exists():
            return candidate, "已复制（同名改名）"
        if candidate.is_file() and _same_file_contents(source, candidate):
            return candidate, "已存在（内容相同）"
    raise OSError(f"无法为同名文件生成可用目标名：{source.name}")


def copy_selected_images(result: LookupResult, destination_root: str | Path) -> CopyResult:
    """Copy all found images into per-patient folders without overwriting files."""

    root = Path(destination_root).resolve()
    root.mkdir(parents=True, exist_ok=True)
    copied: list[CopyRecord] = []
    for image in result.images:
        if not image.image_path:
            continue
        patient_dir = root / _patient_folder_name(image)
        patient_dir.mkdir(parents=True, exist_ok=True)
        for source_text in image.image_path.splitlines():
            source = Path(source_text)
            try:
                if not source.is_file():
                    raise FileNotFoundError(f"源文件不存在：{source}")
                destination, status = _available_destination(source, patient_dir)
                if status != "已存在（内容相同）":
                    shutil.copy2(source, destination)
                copied.append(
                    CopyRecord(
                        query_order=image.query_order,
                        requested_id=image.requested_id,
                        patient_id=image.patient_id,
                        image_filename=image.image_filename,
                        source_path=str(source.resolve()),
                        copy_status=status,
                        destination_path=str(destination.resolve()),
                        error="",
                    )
                )
            except OSError as exc:
                copied.append(
                    CopyRecord(
                        query_order=image.query_order,
                        requested_id=image.requested_id,
                        patient_id=image.patient_id,
                        image_filename=image.image_filename,
                        source_path=str(source),
                        copy_status="拷贝失败",
                        destination_path="",
                        error=str(exc),
                    )
                )
    return CopyResult(destination_root=root, records=tuple(copied))


TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTLE_FILL = PatternFill("solid", fgColor="D9EAF7")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
AMBER_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_GRAY = Side(style="thin", color="D9E2F3")


def _as_excel_date(value: str) -> date | str:
    if not value:
        return ""
    try:
        return date.fromisoformat(value)
    except ValueError:
        return value


def _style_title(sheet, title: str, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, title)
    cell.fill = TITLE_FILL
    cell.font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.sheet_view.showGridLines = False


def _style_header(sheet, row: int, column_count: int) -> None:
    for cell in sheet[row][:column_count]:
        cell.fill = HEADER_FILL
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[row].height = 30


def _set_text_cell(cell, value: object) -> None:
    cell.value = "" if value is None else str(value)
    cell.data_type = "s"
    cell.number_format = "@"


def _add_table(sheet, header_row: int, row_count: int, column_count: int, name: str) -> None:
    if row_count <= 0:
        return
    ref = f"A{header_row}:{get_column_letter(column_count)}{header_row + row_count}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _set_widths(sheet, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_summary_sheet(
    workbook: Workbook, result: LookupResult, copy_result: CopyResult | None
) -> None:
    sheet = workbook.active
    sheet.title = "查询汇总"
    headers = [
        "序号",
        "查询ID",
        "匹配状态",
        "匹配方式",
        "患者ID",
        "姓名",
        "性别",
        "出生日期",
        "图片索引数",
        "已找到图片数",
        "缺失图片数",
        "已复制/已存在",
        "拷贝失败",
        "最近拍摄日期",
        "备注",
    ]
    _style_title(sheet, "TOPCON 数据核查反馈", len(headers))
    sheet["A2"] = "生成时间"
    sheet["B2"] = result.generated_at.replace(tzinfo=None)
    sheet["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    sheet["D2"] = "数据目录"
    _set_text_cell(sheet["E2"], result.dataset.root)
    sheet.merge_cells("E2:O2")

    stats = result.stats
    copy_stats = copy_result.stats if copy_result else {
        "copied": 0,
        "already_present": 0,
        "renamed_conflicts": 0,
        "failed": 0,
    }
    cards = [
        ("查询ID数", stats["requested_ids"]),
        ("匹配患者", stats["matched_patients"]),
        ("含图片索引", stats["patients_with_image_index"]),
        ("实际图片找到", stats["image_files_found"]),
        (
            "复制成功/已存在",
            copy_stats["copied"]
            + copy_stats["already_present"]
            + copy_stats["renamed_conflicts"],
        ),
        ("需关注查询", stats["queries_requiring_attention"]),
    ]
    for index, (label, value) in enumerate(cards):
        column = 1 + index * 2
        label_cell = sheet.cell(4, column, label)
        value_cell = sheet.cell(4, column + 1, value)
        label_cell.fill = SUBTLE_FILL
        label_cell.font = Font(name="Microsoft YaHei", bold=True, color="17365D")
        value_cell.fill = SUBTLE_FILL
        value_cell.font = Font(name="Microsoft YaHei", size=14, bold=True, color="17365D")
        value_cell.number_format = "#,##0"

    header_row = 6
    for column, header in enumerate(headers, start=1):
        sheet.cell(header_row, column, header)
    _style_header(sheet, header_row, len(headers))
    copies_by_query: dict[int, list[CopyRecord]] = defaultdict(list)
    if copy_result:
        for record in copy_result.records:
            copies_by_query[record.query_order].append(record)
    text_columns = {2, 3, 4, 5, 6, 7, 15}
    date_columns = {8, 14}
    for row_index, summary in enumerate(result.summaries, start=header_row + 1):
        values = [
            summary.query_order,
            summary.requested_id,
            summary.match_status,
            summary.match_mode,
            summary.patient_id,
            summary.patient_name,
            summary.sex,
            summary.birth_date,
            summary.image_index_count,
            summary.image_files_found,
            summary.image_files_missing,
            sum(
                row.copy_status
                in {"已复制", "已存在（内容相同）", "已复制（同名改名）"}
                for row in copies_by_query[summary.query_order]
            ),
            sum(
                row.copy_status == "拷贝失败"
                for row in copies_by_query[summary.query_order]
            ),
            summary.latest_capture_date,
            summary.notes,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column)
            if column in text_columns:
                _set_text_cell(cell, value)
            elif column in date_columns:
                cell.value = _as_excel_date(str(value))
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=column == 15)

    _add_table(sheet, header_row, len(result.summaries), len(headers), "QuerySummaryTable")
    first_data = header_row + 1
    last_data = header_row + len(result.summaries)
    if last_data >= first_data:
        status_range = f"C{first_data}:C{last_data}"
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$C{first_data}="{STATUS_COMPLETE}"'], fill=GREEN_FILL),
        )
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'OR($C{first_data}="{STATUS_NOT_FOUND}",$C{first_data}="{STATUS_AMBIGUOUS}")'],
                fill=RED_FILL,
            ),
        )
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'AND($C{first_data}<>"{STATUS_COMPLETE}",$C{first_data}<>"{STATUS_NOT_FOUND}",$C{first_data}<>"{STATUS_AMBIGUOUS}")'],
                fill=AMBER_FILL,
            ),
        )
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = f"A{header_row}:O{max(header_row, last_data)}"
    _set_widths(
        sheet,
        [8, 18, 22, 22, 18, 14, 10, 14, 12, 14, 14, 14, 12, 14, 34],
    )
    sheet.print_title_rows = f"1:{header_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0


def _write_details_sheet(
    workbook: Workbook, result: LookupResult, copy_result: CopyResult | None
) -> None:
    sheet = workbook.create_sheet("图片明细")
    headers = [
        "查询序号",
        "查询ID",
        "匹配方式",
        "患者ID",
        "姓名",
        "性别",
        "出生日期",
        "登记日期",
        "拍摄日期",
        "拍摄时间",
        "采集类型",
        "检查内序号",
        "图片文件名",
        "文件状态",
        "图片完整路径",
        "拷贝状态",
        "拷贝目标路径",
        "HIST记录序号",
    ]
    _style_title(sheet, "图片明细", len(headers))
    sheet["A2"] = "提示"
    sheet["B2"] = "双击界面中的图片明细可打开已找到的文件；Excel 中也可点击路径。"
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(headers))
    header_row = 4
    for column, header in enumerate(headers, start=1):
        sheet.cell(header_row, column, header)
    _style_header(sheet, header_row, len(headers))
    copies_by_image: dict[tuple[int, str], list[CopyRecord]] = defaultdict(list)
    if copy_result:
        for record in copy_result.records:
            copies_by_image[(record.query_order, record.image_filename)].append(record)
    text_columns = {2, 3, 4, 5, 6, 10, 11, 12, 13, 14, 15, 16, 17}
    date_columns = {7, 8, 9}
    for row_index, image in enumerate(result.images, start=header_row + 1):
        copy_records = copies_by_image[(image.query_order, image.image_filename)]
        values = [
            image.query_order,
            image.requested_id,
            image.match_mode,
            image.patient_id,
            image.patient_name,
            image.sex,
            image.birth_date,
            image.registered_date,
            image.capture_date,
            image.capture_time,
            image.capture_source,
            image.image_number_in_exam,
            image.image_filename,
            image.file_status,
            image.image_path,
            "\n".join(row.copy_status for row in copy_records),
            "\n".join(row.destination_path for row in copy_records if row.destination_path),
            image.record_index,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column)
            if column in text_columns:
                _set_text_cell(cell, value)
            elif column in date_columns:
                cell.value = _as_excel_date(str(value))
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=column in {15, 16, 17})
        if image.file_status == FILE_FOUND and image.image_path:
            path_cell = sheet.cell(row_index, 15)
            path_cell.hyperlink = Path(image.image_path).resolve().as_uri()
            path_cell.style = "Hyperlink"
        if len(copy_records) == 1 and copy_records[0].destination_path:
            copy_cell = sheet.cell(row_index, 17)
            copy_cell.hyperlink = Path(copy_records[0].destination_path).resolve().as_uri()
            copy_cell.style = "Hyperlink"

    _add_table(sheet, header_row, len(result.images), len(headers), "ImageDetailsTable")
    first_data = header_row + 1
    last_data = header_row + len(result.images)
    if last_data >= first_data:
        status_range = f"N{first_data}:N{last_data}"
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$N{first_data}="{FILE_FOUND}"'], fill=GREEN_FILL),
        )
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$N{first_data}="{FILE_MISSING}"'], fill=RED_FILL),
        )
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$N{first_data}="{FILE_DUPLICATE}"'], fill=AMBER_FILL),
        )
    sheet.freeze_panes = "A5"
    _set_widths(
        sheet,
        [10, 18, 22, 18, 14, 10, 14, 14, 14, 12, 16, 12, 18, 16, 52, 22, 52, 14],
    )
    sheet.print_title_rows = f"1:{header_row}"
    sheet.page_setup.orientation = "landscape"


def _build_exception_rows(
    result: LookupResult, copy_result: CopyResult | None
) -> list[list[object]]:
    rows: list[list[object]] = []
    for summary in result.summaries:
        if summary.match_status == STATUS_COMPLETE:
            continue
        if summary.match_status in {STATUS_NOT_FOUND, STATUS_AMBIGUOUS}:
            suggestion = "核对患者号及前导零；必要时启用唯一前导零匹配"
        elif summary.match_status == STATUS_NO_INDEX:
            suggestion = "患者清单存在，但 HIST0010 没有该患者的图片索引"
        else:
            suggestion = "检查图片文件夹是否完整拷贝，并核对图片根目录"
        rows.append(
            [
                "查询",
                summary.requested_id,
                summary.patient_id,
                "",
                summary.match_status,
                suggestion,
            ]
        )
    for image in result.images:
        if image.file_status == FILE_FOUND:
            continue
        rows.append(
            [
                "图片",
                image.requested_id,
                image.patient_id,
                image.image_filename,
                image.file_status,
                "补拷图片文件" if image.file_status == FILE_MISSING else "核对重名文件所在目录",
            ]
        )
    if copy_result:
        for record in copy_result.records:
            if record.copy_status != "拷贝失败":
                continue
            rows.append(
                [
                    "拷贝",
                    record.requested_id,
                    record.patient_id,
                    record.image_filename,
                    record.copy_status,
                    record.error,
                ]
            )
    return rows


def _write_exceptions_sheet(
    workbook: Workbook, result: LookupResult, copy_result: CopyResult | None
) -> None:
    sheet = workbook.create_sheet("异常与缺失")
    headers = ["类别", "查询ID", "患者ID", "图片文件名", "状态", "建议处理"]
    _style_title(sheet, "异常与缺失", len(headers))
    header_row = 3
    for column, header in enumerate(headers, start=1):
        sheet.cell(header_row, column, header)
    _style_header(sheet, header_row, len(headers))
    rows = _build_exception_rows(result, copy_result)
    for row_index, values in enumerate(rows, start=header_row + 1):
        for column, value in enumerate(values, start=1):
            _set_text_cell(sheet.cell(row_index, column), value)
        sheet.cell(row_index, 5).fill = RED_FILL if values[4] in {STATUS_NOT_FOUND, STATUS_AMBIGUOUS, FILE_MISSING} else AMBER_FILL
    if not rows:
        sheet.cell(header_row + 1, 1, "未发现异常或缺失")
        sheet.merge_cells(
            start_row=header_row + 1,
            start_column=1,
            end_row=header_row + 1,
            end_column=len(headers),
        )
        sheet.cell(header_row + 1, 1).fill = GREEN_FILL
    _add_table(sheet, header_row, len(rows), len(headers), "ExceptionsTable")
    sheet.freeze_panes = "A4"
    _set_widths(sheet, [12, 18, 18, 20, 24, 46])


def _write_copy_sheet(workbook: Workbook, copy_result: CopyResult | None) -> None:
    sheet = workbook.create_sheet("拷贝记录")
    headers = [
        "查询序号",
        "查询ID",
        "患者ID",
        "图片文件名",
        "源文件路径",
        "拷贝状态",
        "目标文件路径",
        "错误信息",
    ]
    _style_title(sheet, "图片拷贝记录", len(headers))
    sheet["A2"] = "目标目录"
    if copy_result:
        _set_text_cell(sheet["B2"], copy_result.destination_root)
    else:
        sheet["B2"] = "本次未执行图片拷贝"
    sheet.merge_cells("B2:H2")
    header_row = 4
    for column, header in enumerate(headers, start=1):
        sheet.cell(header_row, column, header)
    _style_header(sheet, header_row, len(headers))
    records = copy_result.records if copy_result else ()
    for row_index, record in enumerate(records, start=header_row + 1):
        values = [
            record.query_order,
            record.requested_id,
            record.patient_id,
            record.image_filename,
            record.source_path,
            record.copy_status,
            record.destination_path,
            record.error,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column)
            if column == 1:
                cell.value = value
            else:
                _set_text_cell(cell, value)
            cell.alignment = Alignment(vertical="top", wrap_text=column in {5, 7, 8})
        if record.source_path:
            source_cell = sheet.cell(row_index, 5)
            source_cell.hyperlink = Path(record.source_path).resolve().as_uri()
            source_cell.style = "Hyperlink"
        if record.destination_path:
            target_cell = sheet.cell(row_index, 7)
            target_cell.hyperlink = Path(record.destination_path).resolve().as_uri()
            target_cell.style = "Hyperlink"
        status_cell = sheet.cell(row_index, 6)
        status_cell.fill = RED_FILL if record.copy_status == "拷贝失败" else GREEN_FILL
    _add_table(sheet, header_row, len(records), len(headers), "CopyLogTable")
    sheet.freeze_panes = "A5"
    _set_widths(sheet, [10, 18, 18, 20, 52, 22, 52, 42])


def write_feedback_workbook(
    result: LookupResult,
    path: str | Path,
    *,
    copy_result: CopyResult | None = None,
) -> Path:
    """Write a formatted, auditable Excel feedback workbook."""

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.properties.title = "TOPCON 数据核查反馈"
    workbook.properties.subject = "按患者 ID 核查患者信息、图片索引及实际图片文件"
    workbook.properties.creator = "TOPCON 数据核查工具"
    _write_summary_sheet(workbook, result, copy_result)
    _write_details_sheet(workbook, result, copy_result)
    _write_exceptions_sheet(workbook, result, copy_result)
    _write_copy_sheet(workbook, copy_result)
    workbook.save(output_path)
    return output_path


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="按患者 ID 查询 TOPCON 全量拷贝并生成 Excel 反馈")
    parser.add_argument("data_dir", help="全量拷贝目录或包含 HIST0010 的目录")
    parser.add_argument("--ids", help="逗号、空格或换行分隔的患者 ID")
    parser.add_argument("--ids-file", help="TXT、CSV 或 XLSX ID 文件")
    parser.add_argument("--image-root", help="图片根目录；默认与 data_dir 相同")
    parser.add_argument("--copy-to", help="把找到的图片安全复制到此目录（按患者 ID 分文件夹）")
    parser.add_argument("--allow-leading-zero-match", action="store_true", help="允许忽略前导零的唯一匹配")
    parser.add_argument("--output", default="output/核查反馈.xlsx", help="输出 XLSX 路径")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_argument_parser().parse_args(argv)
    requested: list[str] = []
    if args.ids:
        requested.extend(parse_id_text(args.ids))
    if args.ids_file:
        requested.extend(load_ids_file(args.ids_file))
    requested = normalize_ids(requested)
    if not requested:
        print("查询失败：请通过 --ids 或 --ids-file 提供患者 ID", file=sys.stderr)
        return 2
    try:
        result = lookup_patient_ids(
            args.data_dir,
            requested,
            image_root=args.image_root,
            allow_leading_zero_match=args.allow_leading_zero_match,
        )
        copy_result = (
            copy_selected_images(result, args.copy_to) if args.copy_to else None
        )
        output_path = write_feedback_workbook(
            result, args.output, copy_result=copy_result
        )
    except (OSError, ValueError, TopconFormatError) as exc:
        print(f"查询失败：{exc}", file=sys.stderr)
        return 1
    print(
        json.dumps(
            {
                **result.stats,
                "copy": copy_result.stats if copy_result else None,
                "output": str(output_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
