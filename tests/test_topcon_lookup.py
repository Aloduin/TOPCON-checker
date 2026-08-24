import struct
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from topcon_lookup import (
    FILE_FOUND,
    STATUS_COMPLETE,
    STATUS_NOT_FOUND,
    copy_selected_images,
    load_ids_file,
    lookup_patient_ids,
    write_feedback_workbook,
)


def _put(record: bytearray, start: int, end: int, value: str, encoding="ascii") -> None:
    raw = value.encode(encoding)
    if len(raw) >= end - start:
        raise ValueError("fixture field too long")
    record[start : start + len(raw)] = raw


def _history_record(patient_id="000123", filename="IM000042.JPG") -> bytes:
    record = bytearray(820)
    record[0:4] = b"*IM\0"
    _put(record, 4, 24, patient_id)
    _put(record, 24, 48, "测试患者", "gb18030")
    _put(record, 48, 76, "8岁", "gb18030")
    _put(record, 76, 80, "F")
    _put(record, 80, 132, "2016-01-02")
    _put(record, 132, 160, "2024-03-04")
    _put(record, 160, 188, "2025-06-07")
    _put(record, 620, 640, patient_id)
    _put(record, 640, 652, "2025-06-07")
    _put(record, 652, 664, "08:09:10")
    _put(record, 664, 684, "DC3Capture")
    _put(record, 684, 756, "#1")
    _put(record, 756, 772, filename)
    struct.pack_into("<I", record, 776, 12345)
    struct.pack_into("<I", record, 790, 1)
    return bytes(record)


def _make_dataset(root: Path, *, include_image=True) -> tuple[Path, Path | None]:
    data_dir = root / "hospital_copy"
    data_dir.mkdir()
    (data_dir / "LABEL").write_bytes(b"00010 TOPCON\r\n")
    (data_dir / "patients.TXT").write_bytes(
        "000123,测试患者,8岁,2016-01-02,Female,2024-03-04\r\n".encode(
            "gb18030"
        )
    )
    (data_dir / "HIST0010").write_bytes(_history_record())
    image_path = None
    if include_image:
        image_dir = data_dir / "images" / "2025"
        image_dir.mkdir(parents=True)
        image_path = image_dir / "IM000042.JPG"
        image_path.write_bytes(b"synthetic-jpeg")
    return data_dir, image_path


class TopconLookupTests(unittest.TestCase):
    def test_lookup_finds_metadata_and_actual_image(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            data_dir, image_path = _make_dataset(Path(temp_dir))

            result = lookup_patient_ids(data_dir, ["000123"])

        self.assertEqual(result.summaries[0].match_status, STATUS_COMPLETE)
        self.assertEqual(result.summaries[0].patient_name, "测试患者")
        self.assertEqual(result.summaries[0].image_index_count, 1)
        self.assertEqual(result.images[0].file_status, FILE_FOUND)
        self.assertEqual(result.images[0].image_path, str(image_path.resolve()))

    def test_leading_zero_match_is_disabled_by_default_and_unique_when_enabled(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            data_dir, _ = _make_dataset(Path(temp_dir), include_image=False)

            exact_only = lookup_patient_ids(data_dir, ["123"])
            tolerant = lookup_patient_ids(
                data_dir, ["123"], allow_leading_zero_match=True
            )

        self.assertEqual(exact_only.summaries[0].match_status, STATUS_NOT_FOUND)
        self.assertEqual(tolerant.summaries[0].patient_id, "000123")
        self.assertEqual(tolerant.summaries[0].match_mode, "忽略前导零（唯一匹配）")

    def test_copy_is_idempotent_and_workbook_records_copy(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            data_dir, _ = _make_dataset(root)
            result = lookup_patient_ids(data_dir, ["000123"])
            copy_dir = root / "selected_images"
            first_copy = copy_selected_images(result, copy_dir)
            second_copy = copy_selected_images(result, copy_dir)
            workbook_path = root / "feedback.xlsx"
            write_feedback_workbook(result, workbook_path, copy_result=first_copy)
            workbook = load_workbook(workbook_path, data_only=False)

        self.assertEqual(first_copy.records[0].copy_status, "已复制")
        self.assertEqual(
            Path(first_copy.records[0].destination_path).parent.name,
            "000123_2016-01-02_女",
        )
        self.assertEqual(second_copy.records[0].copy_status, "已存在（内容相同）")
        self.assertEqual(
            workbook.sheetnames,
            ["查询汇总", "图片明细", "异常与缺失", "拷贝记录"],
        )
        self.assertEqual(workbook["查询汇总"]["B7"].value, "000123")
        self.assertEqual(workbook["查询汇总"]["B7"].data_type, "s")
        self.assertEqual(workbook["查询汇总"]["L7"].value, 1)
        self.assertEqual(workbook["拷贝记录"]["F5"].value, "已复制")

    def test_copy_never_overwrites_different_file_with_same_name(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            data_dir, _ = _make_dataset(root)
            result = lookup_patient_ids(data_dir, ["000123"])
            copy_dir = root / "selected_images"
            first_copy = copy_selected_images(result, copy_dir)
            original_target = Path(first_copy.records[0].destination_path)
            self.assertEqual(original_target.parent.name, "000123_2016-01-02_女")
            original_target.write_bytes(b"different-existing-content")

            conflict_copy = copy_selected_images(result, copy_dir)

            renamed_target = Path(conflict_copy.records[0].destination_path)
            self.assertEqual(
                conflict_copy.records[0].copy_status, "已复制（同名改名）"
            )
            self.assertEqual(original_target.read_bytes(), b"different-existing-content")
            self.assertEqual(renamed_target.name, "IM000042__2.JPG")
            self.assertEqual(renamed_target.read_bytes(), b"synthetic-jpeg")

    def test_load_ids_from_excel_preserves_text_identifiers(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "ids.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["患者ID"])
            sheet.append(["000123"])
            sheet.append(["000123"])
            workbook.save(path)

            ids = load_ids_file(path)

        self.assertEqual(ids, ["000123"])


if __name__ == "__main__":
    unittest.main()
